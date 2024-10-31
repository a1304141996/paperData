import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import Calculate_LSM
import os
import config
import text_processing_utils


# 调用Calculate_LSM得到lsm_values
def fetch_lsm_values(text_folder_path, function_word_file):
    try:
        # 调用LSM函数
        lsm_values = Calculate_LSM.LSM(text_folder_path, function_word_file)
        return lsm_values
    except Exception as e:
        # 处理可能发生的异常，例如文件读取错误、计算错误等
        print(f"Error fetching LSM values: {e}")
        return []


# 分段函数：根据空行分割数据
def segment_by_empty_rows(df):
    is_empty_row = df.apply(lambda row: all(pd.isna(cell) or cell == '' for cell in row), axis=1)
    empty_indices = is_empty_row[is_empty_row].index
    empty_indices = np.append(empty_indices, len(df))

    segments = []
    start = 0
    for end in empty_indices:
        segment = df[start:end]
        start = end + 1
        if len(segment) == 0:
            continue
        segments.append(segment)
    return segments


# RLSM 计算函数
def calculate_rlsm(segment):
    function_words = segment.iloc[:, 2]
    function_words = pd.to_numeric(function_words, errors='coerce')

    function_words_b = np.concatenate([function_words[1:].values, [np.nan]])

    rlsm_df = pd.DataFrame({
        'function_words': function_words,
        'function_words_b': function_words_b
    })

    def compute_rlsm(value_a, value_b):
        if pd.notna(value_a) and pd.notna(value_b):
            return 1 - abs(value_a - value_b) / (value_a + value_b + 0.0001)
        else:
            return np.nan

    rlsm_df['rLSM'] = rlsm_df.apply(lambda row: round(compute_rlsm(row['function_words'], row['function_words_b']), 2),
                                    axis=1)

    valid_rlsm = rlsm_df['rLSM'].dropna()

    avg_rlsm = round(valid_rlsm.mean(), 2) if len(valid_rlsm) > 0 else np.nan

    avg_row = pd.DataFrame({
        'function_words': [np.nan],
        'function_words_b': [np.nan],
        'rLSM': [np.nan],
        'M rLSM': [avg_rlsm]
    })

    rlsm_df = pd.concat([avg_row, rlsm_df], ignore_index=True)
    return rlsm_df


# 写入 Excel 文件的函数
def write_to_excel_with_rlsm(xlsx_directory_path, lsm_values):
    # 读取 Excel 文件
    df = pd.read_excel(xlsx_directory_path)

    # 按空行分割数据
    segments = segment_by_empty_rows(df)
    print(len(segments))
    print(len(lsm_values))
    assert len(segments) == len(lsm_values), "每个段落必须有一个对应的LSM值"

    # 初始化用于存储最终合并的列
    df['rLSM'] = np.nan
    df['M rLSM'] = np.nan
    df['LSM'] = np.nan

    # 逐个段落计算 RLSM 并将其写入原始数据框的相应位置
    start_index = 0
    for i, segment in enumerate(segments):
        rlsm_df = calculate_rlsm(segment)
        rlsm_length = len(rlsm_df) - 1  # 减去平均值行

        end_index = start_index + rlsm_length - 1

        if end_index >= start_index:
            # 将 rlsm 和 average_rlsm 列的数据写入原始数据框的相应行
            df.loc[start_index:end_index, 'rLSM'] = rlsm_df.loc[1:, 'rLSM'].values
            df.loc[start_index, 'M rLSM'] = rlsm_df.loc[0, 'M rLSM']
            df.loc[start_index, 'LSM'] = lsm_values[i]  # 将当前段落的LSM值写入

        start_index = end_index + 2  # 跳过空行

    # 使用 openpyxl 直接加载现有文件
    workbook = load_workbook(xlsx_directory_path)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    # 确保列名正确写入
    sheet['D1'] = 'rLSM'
    sheet['E1'] = 'M rLSM'
    sheet['F1'] = 'LSM'

    # 将 Pandas 数据框的结果写回 Excel 文件
    for index, row in df.iterrows():
        sheet[f'D{index + 2}'] = row['rLSM']
        sheet[f'E{index + 2}'] = row['M rLSM']
        sheet[f'F{index + 2}'] = row['LSM']

    # 设置新写入列的宽度
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15

    # 设置所有新列的水平对齐方式，并应用字体格式
    for col in ['D', 'E', 'F']:
        # 设置列标题的格式
        sheet[f'{col}1'].alignment = Alignment(horizontal='center')
        sheet[f'{col}1'].font = Font(bold=True)

        # 设置数据行的对齐方式
        for row in range(2, len(df) + 2):
            cell = sheet[f'{col}{row}']
            cell.alignment = Alignment(horizontal='center')

    # 保存修改后的文件
    workbook.save(xlsx_directory_path)


def process_multiple_xlsx(xlsx_directory_path, text_directory_path, function_word_file):
    for filename in os.listdir(xlsx_directory_path):
        if filename.endswith('.xlsx'):
            xlsx_file_path = os.path.join(xlsx_directory_path, filename)

            # 基于xlsx文件名构造对应的文本文件夹名
            base_filename = os.path.splitext(filename)[0]  # 移除扩展名
            text_folder_path = os.path.join(text_directory_path, base_filename)  # 直接使用同名文件夹
            print(text_folder_path)

            print(f"Processing file: {xlsx_directory_path}\{filename}")
            if os.path.exists(text_folder_path):
                # 如果对应的文件夹存在，则计算LSM值
                lsm_values = fetch_lsm_values(text_folder_path, function_word_file)
                write_to_excel_with_rlsm(xlsx_file_path, lsm_values)
            else:
                print(f"Corresponding text folder not found for {filename}")


process_multiple_xlsx(config.OUTPUT_DIRECTORY, config.SPLIT_DIRECTORY, config.FUNCTION_WORDS_FILE)

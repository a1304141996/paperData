"""
需要对有效咨询划分的逻辑进行去除
"""

import pandas as pd
import os
import openpyxl
import glob
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import text_processing_utils
import config


# 处理文本的每一行，统计词语和功能词
def process_lines(paragraphs, function_words):
    word_counts = []
    func_word_counts = []
    total_func_word_counts = []

    speakers = []  # 用于存储发言者
    statements = []  # 用于存储对话内容
    fw_values = []  # 用于存储FW值

    # paragraph是一个列表，存储了不同的有效咨询段
    for paragraph in paragraphs:
        # lines是一个列表，存储了一个有效咨询段中的全部对话
        lines = paragraph.splitlines()
        # line是一段单一的对话
        for line in lines:
            if ':' not in line:
                continue
            # 去除发送者标识，只保留对话内容
            statement = line.split(':')[1].strip()
            # 获取发言者标识
            speaker = line.split(':')[0].strip().split('->')[0].strip()
            # 存储发言者
            speakers.append(speaker)

            # statement是一个列表，存储了去除发送者标识，只有文字的一个有效咨询的全部对话,用于后续存入xlsx文件
            statements.append(statement)

            # 清理文本，合并为一个连续的句子
            cleaned_sentence = text_processing_utils.clean_and_merge(statement)

            # 处理合并后的句子
            total_func_word_count, total_words, counts, matched_function_words = text_processing_utils.process_sentence(
                cleaned_sentence
                , function_words)

            # 计算FW值
            fw_value = text_processing_utils.calculate_fw_values(total_func_word_count, total_words)
            # 存储FW值
            fw_values.append(round(fw_value * 100, 2))

            # 用于测试，直观查看数据经过处理后的样子，可以删除，不影响流程
            word_counts.append(total_words)  # 该条对话中的词总数
            func_word_counts.append(counts)  # 该条对话中包含的功能词及其个数
            total_func_word_counts.append(total_func_word_count)  # 该条对话中功能此总个数

        # 添加一个空行作为段落间分隔符
        speakers.append('')
        statements.append('')
        fw_values.append('')

    return speakers, statements, fw_values


# 将包含speaker，statement和func_word的数据保存到excel表格中,且文件名与源文件相同，但后缀为‘xlsx’
def save_to_excel(speakers, statements, fw_values, original_filepath, output_folder):
    # 创建DataFrame对象
    data = {
        'Speaker': speakers,
        'Statement': statements,
        'Function Words (%)': fw_values
    }
    df = pd.DataFrame(data)

    # 生成输出文件路径，替换原始文件扩展名为 '.xlsx'
    base_filename = os.path.splitext(os.path.basename(original_filepath))[0]
    output_filepath = os.path.join(output_folder, f'{base_filename}.xlsx')

    # 将DataFrame保存到Excel文件
    df.to_excel(output_filepath, index=False)

    # 由于表格宽度太小，标题部分重叠了，为了不影响整体观感，调整表格宽度以适应标题
    # 打开生成的Excel文件，调整列宽
    workbook = openpyxl.load_workbook(output_filepath)
    worksheet = workbook.active

    # 设置列宽自适应列名，并开启小字体以适应列宽
    column_widths = {
        'Speaker': 10,
        'Statement': 80,
        'Function Words (%)': 20
    }

    # 遍历DataFrame列名获取excel列字母
    for idx, (col_name, width) in enumerate(column_widths.items(), start=1):
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = width

    # 禁用Statement列的自动换行和缩小字体以适应列宽
    for cell in worksheet['B']:
        cell.alignment = Alignment(wrapText=False, shrinkToFit=False)  # wrapText:自动换行；shrinkToFit:缩小字体

    # 保存调整后的excel文件
    workbook.save(output_filepath)


# 读取文件夹当中的所有文本文件，将它们经过处理后保存为xlsx文件并保存在输出文件夹当中
def process_all_files(input_folder, output_folder, function_words):
    # 如果文件夹不存在，则创建它
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 查找文件夹中的所有 txt 文件
    txt_files = glob.glob(os.path.join(input_folder, '*.txt'))

    for txt_file in txt_files:
        paragraphs = text_processing_utils.split_paragraphs(txt_file)
        speakers, statements, fw_values = process_lines(paragraphs, function_words)
        save_to_excel(speakers, statements, fw_values, txt_file, output_folder)
